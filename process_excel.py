from openpyxl import load_workbook

def read_and_print_sheet(excel_file, sheet_name=None):
    """
    Read data from a sheet of an Excel file and print it.
    
    Args:
        excel_file (str): Path to the Excel file
        sheet_name (str): Name of the sheet to read. If None, reads the first sheet.
    """
    # Load the workbook
    workbook = load_workbook(excel_file)
    
    # Show all available sheets
    print(f"Available sheets: {workbook.sheetnames}")
    print("-" * 50)
    
    # Access the specified sheet or the first sheet
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active  # Get the first/active sheet
    
    # Print sheet information
    print(f"Reading from: {excel_file}")
    print(f"Sheet name: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")
    print("-" * 50)
    
    # Iterate through all rows and print the data
    for row in sheet.iter_rows(values_only=True):
        # Print each row as a tuple
        print(row)
    
    # Close the workbook
    workbook.close()

if __name__ == "__main__":
    # Example usage
    # Replace 'your_file.xlsx' with the path to your Excel file
    excel_file_path = '/Users/credd9/Downloads/cleanroom-app1/24.9.2025- SCA - HVAC - Matrix.xlsx'
    
    try:
        read_and_print_sheet(excel_file_path)
    except FileNotFoundError:
        print(f"Error: File '{excel_file_path}' not found.")
        print("Please update the excel_file_path variable with the correct file path.")
    except Exception as e:
        print(f"An error occurred: {e}")
